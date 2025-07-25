import  openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
import os

header = ['Date', 'Description','Category', 'Amount']

#set the directory
os.chdir("C:\\Users\\daram\\pythonscripts")

#define method to create workbook
def createWorkbook():
    workbook = openpyxl.Workbook()
    workbook.create_sheet("Expenses",0)
    worksheet = workbook["Expenses"]
    worksheet.append(header)
    workbook.save("Expenses.xlsx")


def loadWorkbook():
    workbook = openpyxl.load_workbook("Expenses.xlsx")
   # worksheet = workbook["Expenses"]
    return  workbook

def addExpenseToSheet(workbook,  add_expense):
    worksheet = workbook["Expenses"]
    maxrow = worksheet.max_row +1
    print(" worksheet.max_row -> "+ str(worksheet.max_row))
    print(" worksheet.max_column -> "+ str(worksheet.max_column))
    try:
        for row in range(len(add_expense)):
            worksheet[f'A{maxrow}'] = add_expense[0]
            worksheet[f'B{maxrow}'] = add_expense[1]
            worksheet[f'C{maxrow}'] = add_expense[2]
            worksheet[f'D{maxrow}'] = add_expense[3]
        workbook.save("Expenses.xlsx")
    except Exception as error:
        print("error" + str(error))


def viewExpenses(workbook):
    worksheet = workbook["Expenses"]
    for row in worksheet:
        print(row[0].value +" | "+ row[1].value + " | "+ row[2].value + " | " + row[3].value)



option = ["1  - Add Expenses", "2 - View Expenses"]

print(option)


try:
    choice = (input("Enter your choice: "))
    if choice == "1":
        add_expense_to_sheet = []
        print("Add Expenses to Sheet")
        for x in header:
            add_expense_to_sheet.append( input(f" {x}: "))
        try:
            workbook1 = loadWorkbook()
        except Exception as error:
            createWorkbook()
            print("error" + str(error) +" Workbook recreated")

        workbook = loadWorkbook()
        if "Expenses" in workbook.sheetnames:
            addExpenseToSheet(workbook, add_expense_to_sheet)

    elif choice == "2":
        workbook = loadWorkbook()
        viewExpenses(workbook)

except Exception as error:
    print("supply valid number" + str(error))


